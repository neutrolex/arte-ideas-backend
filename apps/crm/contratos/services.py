"""
Servicios del Módulo de Contratos - Arte Ideas CRM
Servicios para exportación y generación de documentos
"""
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.conf import settings
from io import BytesIO
import uuid
from datetime import datetime


class PDFNotImplemented(Exception):
    """Excepción cuando las dependencias de PDF no están instaladas"""
    pass


class ContractPDFService:
    """
    Servicio para generar PDFs de contratos
    """
    
    @staticmethod
    def generate(contrato, template_name='contract'):
        """
        Generar PDF del contrato
        """
        try:
            # Intentar importar weasyprint
            from weasyprint import HTML, CSS
        except ImportError:
            raise PDFNotImplemented("WeasyPrint no está instalado. Instale con: pip install weasyprint")
        
        # Preparar contexto para el template
        context = {
            'contract': {
                'title': contrato.titulo,
                'client_name': contrato.cliente.obtener_nombre_completo(),
                'client': contrato.cliente,
                'contract_type': contrato.tipo_servicio,
                'get_contract_type_display': contrato.get_tipo_servicio_display(),
                'status': contrato.estado,
                'get_status_display': contrato.get_estado_display(),
                'amount': contrato.monto_total,
                'start_date': contrato.fecha_inicio,
                'end_date': contrato.fecha_fin,
                'details': contrato.descripcion,
                'external_ref': contrato.numero_contrato,
                'tenant': contrato.tenant,
            },
            'tenant': contrato.tenant,
            'generated_at': datetime.now(),
            'clausulas': contrato.clausulas.all().order_by('numero_clausula'),
            'pagos': contrato.pagos.all().order_by('-fecha_pago'),
            'total_pagado': sum(pago.monto for pago in contrato.pagos.all()),
            'saldo_pendiente': contrato.saldo_pendiente,
            'porcentaje_adelanto': contrato.porcentaje_adelanto,
        }
        
        # Renderizar template HTML
        template_path = f'exports/{template_name}.html'
        # Si no se especifica template, usar el detallado
        if template_name == 'contract':
            template_path = 'exports/contract_detailed.html'
        
        html_content = render_to_string(template_path, context)
        
        # Generar PDF
        html = HTML(string=html_content)
        pdf_file = html.write_pdf()
        
        # Guardar archivo temporalmente si es necesario
        filename = f"contrato_{contrato.numero_contrato}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        # Actualizar el campo documento_contrato si no existe
        if not contrato.documento_contrato:
            from django.core.files.base import ContentFile
            contrato.documento_contrato.save(
                filename,
                ContentFile(pdf_file),
                save=True
            )
        
        return BytesIO(pdf_file)
    
    @staticmethod
    def generate_response(contrato, filename=None):
        """
        Generar respuesta HTTP con el PDF del contrato
        """
        pdf_content = ContractPDFService.generate(contrato)
        
        if not filename:
            filename = f"contrato_{contrato.numero_contrato}.pdf"
        
        response = HttpResponse(
            pdf_content.getvalue(),
            content_type='application/pdf'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class ContractExcelService:
    """
    Servicio para generar reportes Excel de contratos
    """
    
    @staticmethod
    def generate_contracts_report(contratos, tenant):
        """
        Generar reporte Excel de contratos
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl no está instalado. Instale con: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Contratos"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Headers
        headers = [
            'N° Contrato', 'Cliente', 'Título', 'Tipo Servicio', 'Estado',
            'Fecha Inicio', 'Fecha Fin', 'Monto Total', 'Adelanto', 
            'Saldo Pendiente', '% Adelanto', 'Creado'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Datos
        for row, contrato in enumerate(contratos, 2):
            ws.cell(row=row, column=1, value=contrato.numero_contrato)
            ws.cell(row=row, column=2, value=contrato.cliente.obtener_nombre_completo())
            ws.cell(row=row, column=3, value=contrato.titulo)
            ws.cell(row=row, column=4, value=contrato.get_tipo_servicio_display())
            ws.cell(row=row, column=5, value=contrato.get_estado_display())
            ws.cell(row=row, column=6, value=contrato.fecha_inicio.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=7, value=contrato.fecha_fin.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=8, value=float(contrato.monto_total))
            ws.cell(row=row, column=9, value=float(contrato.adelanto))
            ws.cell(row=row, column=10, value=float(contrato.saldo_pendiente))
            ws.cell(row=row, column=11, value=f"{contrato.porcentaje_adelanto:.1f}%")
            ws.cell(row=row, column=12, value=contrato.creado_en.strftime('%Y-%m-%d %H:%M'))
        
        # Auto-ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Agregar información del tenant
        ws2 = wb.create_sheet("Información")
        ws2.cell(row=1, column=1, value="Estudio Fotográfico:")
        ws2.cell(row=1, column=2, value=tenant.business_name)
        ws2.cell(row=2, column=1, value="Generado:")
        ws2.cell(row=2, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ws2.cell(row=3, column=1, value="Total Contratos:")
        ws2.cell(row=3, column=2, value=len(contratos))
        
        # Guardar en BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    @staticmethod
    def generate_payments_report(contratos, tenant):
        """
        Generar reporte Excel de pagos de contratos
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl no está instalado. Instale con: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Pagos"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Headers
        headers = [
            'N° Contrato', 'Cliente', 'Fecha Pago', 'Monto', 'Método Pago',
            'N° Operación', 'Registrado Por', 'Observaciones'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Datos
        row = 2
        for contrato in contratos:
            for pago in contrato.pagos.all():
                ws.cell(row=row, column=1, value=contrato.numero_contrato)
                ws.cell(row=row, column=2, value=contrato.cliente.obtener_nombre_completo())
                ws.cell(row=row, column=3, value=pago.fecha_pago.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=4, value=float(pago.monto))
                ws.cell(row=row, column=5, value=pago.get_metodo_pago_display())
                ws.cell(row=row, column=6, value=pago.numero_operacion or '')
                ws.cell(row=row, column=7, value=pago.registrado_por.get_full_name())
                ws.cell(row=row, column=8, value=pago.observaciones or '')
                row += 1
        
        # Auto-ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file


class ContractDocumentService:
    """
    Servicio para gestión de documentos de contratos
    """
    
    @staticmethod
    def generate_contract_number(tenant):
        """
        Generar número de contrato único
        """
        from .models import Contrato
        
        # Obtener el último número de contrato del tenant
        last_contract = Contrato.objects.filter(tenant=tenant).order_by('-id').first()
        
        if last_contract:
            # Extraer número del último contrato
            try:
                last_number = int(last_contract.numero_contrato.split('-')[-1])
                new_number = last_number + 1
            except (ValueError, IndexError):
                new_number = 1
        else:
            new_number = 1
        
        # Formato: CT-YYYY-NNNN
        year = datetime.now().year
        return f"CT-{year}-{new_number:04d}"
    
    @staticmethod
    def create_default_clauses(contrato):
        """
        Crear cláusulas por defecto para un contrato
        """
        from .models import ClausulaContrato
        
        default_clauses = [
            {
                'numero_clausula': 1,
                'titulo': 'Objeto del Contrato',
                'contenido': f'El presente contrato tiene por objeto la prestación de servicios de {contrato.get_tipo_servicio_display().lower()} según las especificaciones acordadas.'
            },
            {
                'numero_clausula': 2,
                'titulo': 'Plazo de Ejecución',
                'contenido': f'Los servicios se ejecutarán desde el {contrato.fecha_inicio.strftime("%d/%m/%Y")} hasta el {contrato.fecha_fin.strftime("%d/%m/%Y")}.'
            },
            {
                'numero_clausula': 3,
                'titulo': 'Monto y Forma de Pago',
                'contenido': f'El monto total del contrato es de S/ {contrato.monto_total}. El pago se realizará según los términos acordados.'
            },
            {
                'numero_clausula': 4,
                'titulo': 'Obligaciones del Cliente',
                'contenido': 'El cliente se compromete a proporcionar toda la información y materiales necesarios para la correcta ejecución del servicio.'
            },
            {
                'numero_clausula': 5,
                'titulo': 'Obligaciones del Proveedor',
                'contenido': f'{contrato.tenant.business_name} se compromete a ejecutar los servicios con la calidad y en los plazos acordados.'
            }
        ]
        
        for clause_data in default_clauses:
            ClausulaContrato.objects.create(
                contrato=contrato,
                **clause_data
            )
    
    @staticmethod
    def validate_contract_data(data, tenant):
        """
        Validar datos del contrato antes de crear
        """
        errors = {}
        
        # Validar número de contrato único
        from .models import Contrato
        numero_contrato = data.get('numero_contrato')
        if numero_contrato:
            if Contrato.objects.filter(tenant=tenant, numero_contrato=numero_contrato).exists():
                errors['numero_contrato'] = 'Ya existe un contrato con este número'
        
        # Validar fechas
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')
        if fecha_inicio and fecha_fin and fecha_inicio > fecha_fin:
            errors['fecha_fin'] = 'La fecha de fin debe ser posterior a la fecha de inicio'
        
        # Validar montos
        monto_total = data.get('monto_total', 0)
        adelanto = data.get('adelanto', 0)
        if adelanto > monto_total:
            errors['adelanto'] = 'El adelanto no puede ser mayor al monto total'
        
        return errors